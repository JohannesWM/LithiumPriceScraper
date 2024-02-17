import requests
import bs4
from Scrape_Testing.data_obj import LithiumData, GFEXData
from openpyxl import load_workbook
import datetime

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) \
    AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.114 Safari/537.36'
}

Link = "https://www.metal.com/price/New%20Energy/Lithium"

"""
This function takes a link to a website and returns the html of the page (str format).
Chronologically STEP 1 of the scraping process.
"""


def request_to_site(link):
    page_html = requests.get(link, headers=headers).text

    return page_html


"""
This function will dump a soup obj into a designated html file.
"""


def dump_html(soup_html, file_name):
    if not file_name.endswith(".html"):
        file_name += ".html"

    with open(file_name, "w", encoding="utf-8") as file:
        file.write(soup_html)


"""
This function will take a txt_html (str format) and return a list of LithiumData objects.
Chronologically STEP 2 of the scraping process.
"""


def soup_scrape(txt_html):
    LithiumDataList = []

    try:
        page_soup = bs4.BeautifulSoup(txt_html, "html.parser")

        lithium_rows = page_soup.find_all("div", {"class": "row___1xJWs close___30tSe"})

        for litRow in lithium_rows:
            lit_name, lit_price, lit_date = litRow.find("a", {"class": "link___2BznB item___ku9Fy"}), litRow.find("div", {
                "style": "flex:;width:106px;padding-right:12px;justify-content:flex-end"}), litRow.find("div", {
                "style": "flex:;width:120px;padding-right:24px;justify-content:flex-end"})

            if lit_name and lit_price and lit_date:
                LithiumDataList.append(LithiumData(lit_name.text, lit_price.text, lit_date.text))

        return LithiumDataList
    except Exception as e:
        print(e)
        return -1


"""
This function will dump a list of LithiumData objects into a designated xlsx file.
"""


def dump_xlsx(LithiumData_List, path="LithiumData.xlsx", GFEX=False):

    if LithiumData_List == -1:
        raise Exception("LithiumData_List is empty.")

    cnum_sheet = None
    cnum = "Current Row #"
    must_scrape = False
    start_row = -1

    wb = load_workbook(path)
    ws = wb["Lithium_data"]

    # The algorithm will check if workbook has a sheet called "Current Row #".  If it does, it will check if there
    # is number within the assigned square, if not must_scrape will be set to True.
    try:
        cnum_sheet = wb[cnum]
        start_row = cnum_sheet.cell(row=2, column=1).value

        if isinstance(start_row, int):
            pass
        else:
            must_scrape = True
    except KeyError as e:
        print(e)
        pass

    # if the algorithm doesn't know what row to begin putting items into, it will search for the first empty row
    if must_scrape:
        empty = False
        check_row = 2
        while not empty:

            box_val = ws.cell(row=check_row, column=1).value
            if box_val is None or box_val == "" or box_val == " ":  # if the box is empty
                empty = True
            else:
                check_row += 1

        start_row = check_row

    for row_obj in LithiumData_List:

        if not GFEX:
            ws.cell(row=start_row, column=1, value=str(row_obj.name))
            ws.cell(row=start_row, column=2, value=str(row_obj.price))
            ws.cell(row=start_row, column=3, value=str(row_obj.date))
        elif GFEX:
            ws.cell(row=start_row, column=1, value=str(row_obj.name))
            ws.cell(row=start_row, column=2, value=str(row_obj.price))
            ws.cell(row=start_row, column=3, value=str(row_obj.date))
            ws.cell(row=start_row, column=4, value=str(row_obj.lit_High))
            ws.cell(row=start_row, column=5, value=str(row_obj.lit_Low))
            ws.cell(row=start_row, column=6, value=str(row_obj.volume))
        else:
            raise Exception("GFEX is not a boolean value.")

        start_row += 1

    cnum_sheet.cell(row=2, column=1, value=start_row)

    wb.save(path)
    wb.close()


"""
This function scrapes teh GFEX portion of the website.  Input: page's html
"""


def GFEX_scrape(page_html):

    GFEXDataList = []
    lit_price = ""
    lit_High = ""
    lit_Low = ""
    Volume = ""

    try:
        page_soup = bs4.BeautifulSoup(page_html, "html.parser")

        lithium_rows = page_soup.find_all("div", {"class": "row___1xJWs close___30tSe"})

        for litRow in lithium_rows:
            lit_name = litRow.find("a", {"class": "link___2BznB item___ku9Fy"})

            specific_data = litRow.find_all("div", {"style": "flex:;width:80px;padding-right:12px;justify-content:"
                                                             "flex-end"})
            if specific_data.__len__() == 5:
                lit_price = specific_data[0]
                lit_High = specific_data[1]
                lit_Low = specific_data[2]
                Volume = specific_data[3]

            if lit_name and lit_price and lit_High and lit_Low and Volume:
                GFEXDataList.append(GFEXData(lit_name.text, lit_price.text, str(datetime.date.today()),
                                             lit_High.text, lit_Low.text, Volume.text))

        return GFEXDataList
    except Exception as e:
        print(e)
        return -1
